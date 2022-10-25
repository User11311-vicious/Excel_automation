import pandas as pd
from openpyxl import load_workbook

# Load in the workbook
wb_labor_actions = load_workbook('/home/ldprpc15/Desktop/Imaginarium/Фрезеровщик.xlsx')
wb_big_bloks = load_workbook('/home/ldprpc15/Desktop/Imaginarium/Profstandart_bloks_markers_(frez).xlsx')

sheet_la = wb_labor_actions.active
sheet_1 = wb_big_bloks.active

n = input()

big_block = []
markers = []

for j in range(2, sheet_1.max_row):
    big_block.append(sheet_1.cell(row=j, column=2).value)

for d in range(2, sheet_1.max_row):
    markers.append(sheet_1.cell(row=d, column=3).value)

for j in range(2, sheet_1.max_row-1):
    print(j-1,' - ',big_block[j-2],' - marker - ',markers[j-2])
    


#check for labor actions


ls = []
ls_i = []
blok = []
number_bb = []

for i in range(2, sheet_la.max_row):
    labor_action = sheet_la.cell(row=i, column=2).value
    
    for j in range(len(markers)):
        current_marker = markers[j]
        current_big_blok = big_block[j]
        if current_marker in labor_action:
            print(i-1,' -  ',labor_action, ' - big blok - ', current_big_blok, ' - marker - ', current_marker)
            ls.append(labor_action)
            ls_i.append(i)
            blok.append(current_big_blok)
            number_bb.append(current_marker)

        
        

dmm = {'№': ls_i,
'Трудовые действия' : ls,
'Marker' : number_bb,
'Крупные блоки' : blok}

df = pd.DataFrame.from_dict(dmm, orient='index')
df = df.transpose() 
df.to_excel('/home/ldprpc15/Desktop/Imaginarium/Result/profstandart(MATVEYx).xlsx', sheet_name = 'TД', index=False)


print('programm END')
